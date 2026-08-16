#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
memory-kb 存储记忆同步器（方案A：md 主数据源 + xlsx 派生视图 + watchdog 双向同步）
- md 为唯一事实源；xlsx 为派生视图，人工也可改 xlsx，改后反向同步回 md
- 每主题一对：memory-kb/<主题>.md + memory-kb/<主题>.xlsx
- 总目录：memory-kb/总目录.md + 总目录.xlsx（存储区，缓存不放索引）
- 防循环：记录已写文件的 mtime，忽略自己触发的变更
用法：
  python memorykb_sync.py sync [--kb <目录>]   # 全量：md->xlsx（对缺失/过期 xlsx）
  python memorykb_sync.py syncall [--kb <目录>]  # 强制：所有 md -> xlsx
  python memorykb_sync.py back [--kb <目录>]   # 全量：xlsx->md（反向）
  python memorykb_sync.py index [--kb <目录>]  # 重建总目录（扫描 *.md）
  python memorykb_sync.py watch [--kb <目录>]  # 常驻 watchdog（默认模式）
路径解析优先级：--kb 参数 > 环境变量 MEMORY_KB > 默认 <脚本同目录>/memory-kb
依赖：openpyxl、watchdog（pip install -r requirements.txt）

L1 缓存记忆（MEMORY.md / USER.md）的 Excel 管理入口：
  python memorykb_sync.py sync --l1 MEMORY.md   # L1 md -> xlsx（同目录同名 .xlsx）
  python memorykb_sync.py back --l1 MEMORY.md   # L1 xlsx -> md（改过的行重组，未改行保留原文）
"""
import os, re, sys, time, argparse
from pathlib import Path

KB = None  # 运行时解析：--kb 参数 > 环境变量 MEMORY_KB > 默认 <脚本同目录>/memory-kb
SELF_WRITES = {}  # path -> mtime we wrote

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    sys.exit("缺少 openpyxl：pip install openpyxl")

MD_HEAD_META = re.compile(r"^>\s*(?:标签|meta)\s*[:：]\s*(.+)$", re.I)


# ---------- md -> xlsx ----------
def md_to_xlsx(md_path: Path) -> Path:
    """把 md 逐行映射到 xlsx：A=行类型, B=内容。表格行按管道解析多列。"""
    lines = md_path.read_text(encoding="utf-8").splitlines()
    wb = Workbook()
    ws = wb.active
    ws.title = "content"
    for ln in lines:
        if not ln.strip():
            ws.append(["empty", ""])
        elif ln.startswith("|"):
            cells = [c.strip() for c in ln.strip().strip("|").split("|")]
            ws.append(["table"] + cells)
        elif ln.startswith("## "):
            ws.append(["h2", ln[3:].strip()])
        elif ln.startswith("# "):
            ws.append(["h1", ln[2:].strip()])
        elif ln.startswith("> "):
            ws.append(["quote", ln[2:].strip()])
        elif ln.startswith(("- ", "* ")):
            ws.append(["list", ln[2:].strip()])
        else:
            ws.append(["para", ln.strip()])
    # 列宽
    for col, w in zip("ABCDEFGHIJ", (10, 60, 40, 40, 40, 30, 30, 30, 30, 30)):
        ws.column_dimensions[col].width = w
    xlsx_path = md_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    SELF_WRITES[str(xlsx_path)] = xlsx_path.stat().st_mtime
    return xlsx_path


# ---------- xlsx -> md ----------
def xlsx_to_md(xlsx_path: Path) -> Path:
    wb = load_workbook(xlsx_path, read_only=False)
    ws = wb.active
    out = []
    for row in ws.iter_rows(values_only=True):
        t = row[0] if row and row[0] else "para"
        t = str(t)
        if t == "empty":
            out.append("")
        elif t == "h1":
            out.append("# " + str(row[1]))
        elif t == "h2":
            out.append("## " + str(row[1]))
        elif t == "quote":
            out.append("> " + str(row[1]))
        elif t == "list":
            out.append("- " + str(row[1]))
        elif t == "table":
            cells = [str(c) if c is not None else "" for c in row[1:]]
            out.append("| " + " | ".join(cells) + " |")
        else:
            out.append(str(row[1]) if row[1] is not None else "")
    md_path = xlsx_path.with_suffix(".md")
    md_path.write_text("\n".join(out), encoding="utf-8")
    SELF_WRITES[str(md_path)] = md_path.stat().st_mtime
    return md_path


# ---------- L1 缓存记忆（MEMORY.md / USER.md）Excel 管理 ----------
_L1_TAG_RE = re.compile(r"^\s*\[([^\]]+)\]\s*(.*)$", re.S)   # [标签] 内容
_L1_LIST_RE = re.compile(r"^\s*-\s+(.*)$", re.S)             # - 内容（无标签条目）


def split_l1_blocks(text: str) -> list:
    """把 L1 md 拆成 [(块原文, 块后分隔符), ...]，分隔符含 § 及其周围空白，反写可逐字节还原。
    优先按 § 分隔（协议格式）；无 § 时按顶层 `- ` 列表项拆分（兼容现状，分隔符为 \\n）。"""
    if "§" in text:
        tokens = re.split(r"(\s*§\s*)", text)
        out = []
        for i in range(0, len(tokens) - 1, 2):
            blk = tokens[i].strip()
            if blk:
                out.append((blk, tokens[i + 1] if i + 1 < len(tokens) else ""))
        if tokens and tokens[-1].strip():
            out.append((tokens[-1].strip(), "\n" if text.endswith("\n") else ""))
        return out
    blocks, cur = [], None
    for ln in text.splitlines():
        if ln.strip().startswith("- "):
            if cur is not None:
                blocks.append(("\n".join(cur).strip(), "\n"))
            cur = [ln]
        elif cur is not None:
            cur.append(ln)
    if cur is not None:
        blocks.append(("\n".join(cur).strip(), "\n" if text.endswith("\n") else ""))
    return blocks


def parse_l1_block(block: str) -> tuple:
    """解析一个块 -> (类型, 标签, 内容)。类型：tag=[标签]前缀 / list=-列表 / plain=纯文本。"""
    b = block.strip()
    m = _L1_TAG_RE.match(b)
    if m:
        return "tag", m.group(1).strip(), m.group(2).strip()
    m = _L1_LIST_RE.match(b)
    if m:
        return "list", "", m.group(1).strip()
    return "plain", "", b


def l1_to_xlsx(md_path: Path) -> Path:
    """L1 md -> xlsx：A=序号, B=标签, C=内容, D=原文, E=分隔符（反写还原用，勿改）"""
    text = md_path.read_text(encoding="utf-8")
    wb = Workbook()
    ws = wb.active
    ws.title = "L1"
    ws.append(["序号", "标签", "内容", "原文（只读参考：未修改的行反写时原样保留）", "分隔符（勿改）"])
    for i, (b, sep) in enumerate(split_l1_blocks(text), 1):
        kind, tag, content = parse_l1_block(b)
        ws.append([i, tag, content, b, sep])
    for col, w in zip("ABCDE", (6, 14, 80, 60, 10)):
        ws.column_dimensions[col].width = w
    xlsx_path = md_path.with_suffix(".xlsx")
    wb.save(xlsx_path)
    SELF_WRITES[str(xlsx_path)] = xlsx_path.stat().st_mtime
    return xlsx_path


def xlsx_to_l1(xlsx_path: Path) -> Path:
    """L1 xlsx -> md：改过的行重组（[标签] 内容 / - 列表 / 纯文本）；未改的行保留原文，分隔符原样还原"""
    wb = load_workbook(xlsx_path, read_only=False)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        tag = str(row[1]).strip().strip("[]").strip() if len(row) > 1 and row[1] else ""
        content = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        orig = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        sep = str(row[4]) if len(row) > 4 and row[4] is not None else ""
        if orig:  # 未修改则保留原文（判断：重新解析原文与当前行一致）
            kind0, t0, c0 = parse_l1_block(orig)
            if t0.strip() == tag and c0.strip() == content:
                out.append(orig + sep)
                continue
            if tag:  # 修改过：按新标签重组
                out.append(f"[{tag}] {content}{sep}")
            elif kind0 == "list":  # 原来是列表项
                out.append(f"- {content}{sep}")
            else:  # 纯文本块
                out.append(content + sep)
        else:  # 原文列被清空 / 新增行：按当前行列重组，未填分隔符则默认换行
            raw_sep = sep
            sep = sep or "\n"
            piece = f"[{tag}] {content}" if tag else content
            if out and not out[-1].endswith("\n") and not raw_sep.startswith("\n"):
                piece = "\n" + piece  # 前置换行（单换行），避免与上一块粘连
            out.append(piece + sep)
    md_path = xlsx_path.with_suffix(".md")
    md_path.write_text("".join(out), encoding="utf-8")
    SELF_WRITES[str(md_path)] = md_path.stat().st_mtime
    return md_path


# ---------- 判定与同步 ----------
def _is_self(path: str) -> bool:
    try:
        return SELF_WRITES.get(path) == Path(path).stat().st_mtime
    except OSError:
        return False


def sync_file(path: Path, direction: str = "auto"):
    """auto: 以较新者为源；md: 强制 md->xlsx；xlsx: 强制 xlsx->md"""
    s = str(path)
    if _is_self(s):
        return None
    if path.suffix == ".md" and direction in ("auto", "md"):
        xp = path.with_suffix(".xlsx")
        if not xp.exists() or path.stat().st_mtime > xp.stat().st_mtime:
            return md_to_xlsx(path)
    elif path.suffix == ".xlsx" and direction in ("auto", "xlsx"):
        mp = path.with_suffix(".md")
        if mp.exists() and path.stat().st_mtime > mp.stat().st_mtime:
            return xlsx_to_md(path)
    return None


def sync_all(direction="auto", force=False):
    changed = []
    for p in sorted(KB.glob("*.md")):
        if p.name == "总目录.md":
            continue
        xp = p.with_suffix(".xlsx")
        if force or not xp.exists() or p.stat().st_mtime > xp.stat().st_mtime:
            changed.append(str(md_to_xlsx(p)))
    if not force:  # force 时已由 md 全量生成，反向检查会重复覆盖
        for p in sorted(KB.glob("*.xlsx")):
            mp = p.with_suffix(".md")
            if mp.exists() and p.name != "总目录.xlsx" and p.stat().st_mtime > mp.stat().st_mtime:
                changed.append(str(xlsx_to_md(p)))
    return changed


# ---------- 总目录 ----------
def _meta_of(md_path: Path) -> str:
    for ln in md_path.read_text(encoding="utf-8").splitlines()[:6]:
        m = MD_HEAD_META.match(ln)
        if m:
            return m.group(1).strip()
    return ""


def _summary_of(md_path: Path) -> str:
    """取第一个普通段落或第一个 h2 下首句作摘要"""
    txt = md_path.read_text(encoding="utf-8").splitlines()
    for ln in txt[1:]:
        s = ln.strip()
        if s and not s.startswith((">", "#", "|", "-")):
            return s[:40]
    return ""


def build_index():
    rows = [["文件名", "一级", "二级", "三级", "一句话摘要", "修改日期"]]
    for p in sorted(KB.glob("*.md")):
        if p.name == "总目录.md":
            continue
        meta = _meta_of(p)
        parts = [x.strip() for x in meta.split("/")] if meta else ["", "", ""]
        parts += [""] * (3 - len(parts))
        st = time.strftime("%Y-%m-%d", time.localtime(p.stat().st_mtime))
        rows.append([p.name, parts[0], parts[1], parts[2], _summary_of(p), st])
    # 写 md
    md_lines = ["# 存储记忆总目录", "", "> 本文件由 memorykb_sync.py index 自动生成，勿手改。", ""]
    md_lines.append("| 文件名 | 一级 | 二级 | 三级 | 一句话摘要 | 修改日期 |")
    md_lines.append("|---|---|---|---|---|---|")
    for r in rows[1:]:
        md_lines.append("| " + " | ".join(str(x) for x in r) + " |")
    (KB / "总目录.md").write_text("\n".join(md_lines), encoding="utf-8")
    # 写 xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "总目录"
    for r in rows:
        ws.append(r)
    for col, w in zip("ABCDEF", (22, 12, 14, 24, 46, 12)):
        ws.column_dimensions[col].width = w
    wb.save(KB / "总目录.xlsx")
    return rows


# ---------- watchdog ----------
def watch():
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler

    class H(FileSystemEventHandler):
        def on_modified(self, ev):
            if ev.is_directory:
                return
            try:
                sync_file(Path(ev.src_path))
            except Exception as e:
                print(f"[sync] {ev.src_path} -> ERR {e}")

        def on_created(self, ev):
            if ev.is_directory:
                return
            try:
                sync_file(Path(ev.src_path))
            except Exception as e:
                print(f"[sync] {ev.src_path} -> ERR {e}")

    obs = Observer()
    obs.schedule(H(), str(KB), recursive=False)
    obs.start()
    print(f"[watch] memory-kb 监听中：{KB}（Ctrl+C 退出）")
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        obs.stop()
    obs.join()


def resolve_kb(args_kb: str | None) -> Path:
    """解析 memory-kb 路径：--kb 参数 > 环境变量 MEMORY_KB > 默认 <脚本同目录>/memory-kb"""
    global KB
    if args_kb:
        kb = Path(args_kb)
    elif os.environ.get("MEMORY_KB"):
        kb = Path(os.environ["MEMORY_KB"])
    else:
        kb = Path(__file__).resolve().parent.parent / "memory-kb"
    KB = kb
    return kb


def main():
    ap = argparse.ArgumentParser(description="memory-kb 存储记忆同步器（含 L1 缓存记忆 Excel 管理）")
    ap.add_argument("cmd", nargs="?", default="watch", choices=["sync", "syncall", "back", "index", "watch"])
    ap.add_argument("--kb", help="memory-kb 目录路径（默认：环境变量 MEMORY_KB，再默认脚本同目录的 memory-kb/）")
    ap.add_argument("--l1", help="L1 缓存记忆文件路径（MEMORY.md/USER.md）；配合 sync/syncall/back 使用，把该文件导出/反写 Excel")
    ap.add_argument("--force", action="store_true")
    a = ap.parse_args()

    # ---- L1 模式：直接处理单个缓存记忆文件，不走 KB 目录 ----
    if a.l1:
        p = Path(a.l1)
        if not p.exists():
            sys.exit(f"错误：--l1 文件不存在：{p}")
        if a.cmd in ("sync", "syncall"):
            print("L1 导出 Excel：", l1_to_xlsx(p))
        elif a.cmd == "back":
            xp = p.with_suffix(".xlsx")
            if not xp.exists():
                sys.exit(f"错误：找不到 {xp}（先运行 sync --l1 生成）")
            print("L1 从 Excel 反写：", xlsx_to_l1(xp))
        else:
            sys.exit("--l1 仅支持 sync / syncall / back")
        return

    kb = resolve_kb(a.kb)
    if not kb.exists():
        kb.mkdir(parents=True, exist_ok=True)
        print(f"[init] 已创建 memory-kb 目录：{kb}")
    if not kb.is_dir():
        sys.exit(f"错误：--kb 指向的不是目录：{kb}")

    if a.cmd == "sync":
        print("全量同步：", sync_all())
    elif a.cmd == "syncall":
        print("强制重建：", sync_all(force=True))
    elif a.cmd == "back":
        print("反向同步：", sync_all("xlsx"))
    elif a.cmd == "index":
        rows = build_index()
        print(f"总目录已重建：{len(rows)-1} 个条目 -> {kb}")
    else:
        sync_all()  # 启动时先补一次
        watch()


if __name__ == "__main__":
    main()
